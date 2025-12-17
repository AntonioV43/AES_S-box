from openpyxl import Workbook
from openpyxl.styles import Font


def export_to_excel(
    filename,
    affine_matrix,
    key_hex,
    plaintext,
    ciphertext_hex,
    decrypted,
    sbox,
    enc_trace,
    dec_trace,
    crypto_result
):
    wb = Workbook()
    bold = Font(bold=True)

    # ================= Sheet 1 : Metadata =================
    ws = wb.active
    ws.title = "Metadata"

    meta = [
        ("Affine Matrix", affine_matrix),
        ("AES Key (HEX)", key_hex),
        ("Plaintext", plaintext),
        ("Ciphertext (HEX)", ciphertext_hex),
        ("Decrypted Plaintext", decrypted),
    ]

    for i, (k, v) in enumerate(meta, start=1):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = bold
        ws[f"B{i}"] = v

    # ================= Sheet 2 : S-box =================
    ws = wb.create_sheet("S-box")

    ws["A1"] = "S-box"
    ws["A1"].font = bold

    for i in range(16):
        ws.cell(row=2, column=i+2, value=f"{i:X}").font = bold
        ws.cell(row=i+3, column=1, value=f"{i:X}").font = bold

    for r in range(16):
        for c in range(16):
            ws.cell(row=r+3, column=c+2, value=f"{sbox[r*16+c]:02X}")

    # ================= Sheet 3 : Encrypt Trace =================
    ws = wb.create_sheet("Encrypt Trace")
    row = 1

    for rnd in enc_trace:
        ws[f"A{row}"] = f"Round {rnd['round']}"
        ws[f"A{row}"].font = bold
        row += 1

        for step, state in rnd.items():
            if step == "round":
                continue

            ws[f"A{row}"] = step
            ws[f"A{row}"].font = bold
            row += 1

            for r in state:
                for c, v in enumerate(r):
                    ws.cell(row=row, column=c+1, value=v)
                row += 1
            row += 1
        row += 1

    # ================= Sheet 4 : Decrypt Trace =================
    ws = wb.create_sheet("Decrypt Trace")
    row = 1

    for rnd in dec_trace:
        ws[f"A{row}"] = f"Round {rnd['round']}"
        ws[f"A{row}"].font = bold
        row += 1

        for step, state in rnd.items():
            if step == "round":
                continue

            ws[f"A{row}"] = step
            ws[f"A{row}"].font = bold
            row += 1

            for r in state:
                for c, v in enumerate(r):
                    ws.cell(row=row, column=c+1, value=v)
                row += 1
            row += 1
        row += 1

    # ================= Sheet 5 : Crypto Tests =================
    ws = wb.create_sheet("Crypto Tests")

    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A1"].font = ws["B1"].font = bold

    for i, (k, v) in enumerate(crypto_result.items(), start=2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    wb.save(filename)
